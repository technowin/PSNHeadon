import json
import pydoc
from django.contrib import messages
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render,redirect
from django.contrib.auth import authenticate, login ,logout,get_user_model
from grpc import Status
from Account.forms import RegistrationForm
from Account.models import *
from Masters.models import *
from datetime import date
from Masters.models import site_master as sit
from Masters.models import SlotDetails as slot
from Masters.models import company_master as com
from Payroll.models import * 
import Db 
import re
from datetime import datetime

import bcrypt
from django.contrib.auth.decorators import login_required
from Masters.serializers import * 
from Notification.models import notification_log
from Notification.serializers import NotificationSerializer
from PSNHeadon.encryption import *
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph
from Account.utils import decrypt_email, encrypt_email
import requests
import traceback
import pandas as pd
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, Border, Side
import calendar
from datetime import datetime, timedelta
from django.utils import timezone
from datetime import timedelta
from django.db.models import Q, Count

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.tokens import AccessToken
from django.utils import timezone
from .models import Log, sc_roster, sc_employee_master, CustomUser 
from django.shortcuts import render
from django.db import connection

@login_required
def masters(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    pre_url = request.META.get('HTTP_REFERER')
    header, data = [], []
    entity, type, name = '', '', ''
    global user
    user  = request.session.get('user_id', '')
    try:
         
        if request.method=="GET":
            entity = request.GET.get('entity', '')
            type = request.GET.get('type', '')
            cursor.callproc("stp_get_masters",[entity,type,'name',user])
            for result in cursor.stored_results():
                datalist1 = list(result.fetchall())
            name = datalist1[0][0]
            cursor.callproc("stp_get_masters", [entity, type, 'header',user])
            for result in cursor.stored_results():
                header = list(result.fetchall())
            cursor.callproc("stp_get_masters",[entity,type,'data',user])
            for result in cursor.stored_results():
                if (entity == 'em' or entity == 'sm' or entity == 'cm' or entity == 'menu' or entity == 'user' or entity =='sd' or entity =='dm' or entity =='um') and type !='err': 
                    data = []
                    rows = result.fetchall()
                    for row in rows:
                        encrypted_id = encrypt_parameter(str(row[0]))
                        data.append((encrypted_id,) + row[1:])
                else: data = list(result.fetchall())

            cursor.callproc("stp_get_dropdown_values",['company'])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())
            cursor.callproc("stp_get_dropdown_values",['site'])
            for result in cursor.stored_results():
                site_name = list(result.fetchall())
            cursor.callproc("stp_get_dropdown_values",['designation'])
            for result in cursor.stored_results():
                designation_name = list(result.fetchall())
            
            if entity == 'r' and type == 'i':
                cursor.callproc("stp_get_assigned_company",[user])
                for result in cursor.stored_results():
                    company_names = list(result.fetchall())
            if entity == 'r' and type == 'ed':
                month_year =str(request.GET.get('month', ''))
                if month_year == '':
                    year,month = '',''
                else: year,month = month_year.split('-')
                employee_id = request.GET.get('empid', '')
                cursor.callproc("stp_get_edit_roster",[employee_id,month,year,'1'])
                for result in cursor.stored_results():
                    data = list(result.fetchall())
                cursor.callproc("stp_get_edit_roster",[employee_id,month,year,'2'])
                for result in cursor.stored_results():
                    header = list(result.fetchall())
            if entity == 'urm' and (type == 'acu' or type == 'acr'):
                cursor.callproc("stp_get_access_control",[entity,type])
                for result in cursor.stored_results():
                    header = list(result.fetchall())
                cursor.callproc("stp_get_access_control",[entity,'comp'])
                for result in cursor.stored_results():
                    company_names = list(result.fetchall())
                cursor.callproc("stp_get_access_control",[entity,'site'])
                for result in cursor.stored_results():
                    data = list(result.fetchall())
                
        if request.method=="POST":
            entity = request.POST.get('entity', '')
            type = request.POST.get('type', '')
            if entity == 'r' and type == 'ed':
                ids = request.POST.getlist('ids[]', '')
                shifts = request.POST.getlist('shifts[]', '')
                for id,shift in zip(ids, shifts):
                    cursor.callproc("stp_post_roster",[id,shift])
                    for result in cursor.stored_results():
                        datalist = list(result.fetchall())
                if datalist[0][0] == "success":
                    messages.success(request, 'Data updated successfully !')
            if entity == 'urm' and (type == 'acu' or type == 'acr'):
                
            #     created_by = request.session.get('user_id', '')
            #     ur = request.POST.get('ur', '')
            #     selected_company_ids = list(map(int, request.POST.getlist('company_id')))
            #     selected_worksites  = request.POST.getlist('worksite')
            #     company_worksite_map = {}
                
            #     if not selected_company_ids or not selected_worksites:
            #         messages.error(request, 'Company or worksite data is missing!')
            #         return redirect(f'/masters?entity={entity}&type=urm')
            #     if type not in ['acu', 'acr'] or not ur:
            #         messages.error(request, 'Invalid data received.')
            #         return redirect(f'/masters?entity={entity}&type=urm')
                
            #     cursor.callproc("stp_get_company_worksite",[",".join(request.POST.getlist('company_id'))])
            #     for result in cursor.stored_results():
            #         company_worksites  = list(result.fetchall())
                    
            #     for company_id, worksite_name in company_worksites:
            #         if company_id not in company_worksite_map:
            #             company_worksite_map[company_id] = []
            #         company_worksite_map[company_id].append(worksite_name)
                
            #     filtered_combinations = []
            #     for company_id in selected_company_ids:
            #         valid_worksites = company_worksite_map.get(company_id, [])
            #         # Filter worksites that were actually selected by the user
            #         selected_valid_worksites = [ws for ws in selected_worksites if ws in valid_worksites]
            #         filtered_combinations.extend([(company_id, ws) for ws in selected_valid_worksites])
                    
            #     cursor.callproc("stp_delete_access_control",[type,ur])
            #     r=''
            #     for company_id, worksite in filtered_combinations:
            #         cursor.callproc("stp_post_access_control",[type,ur,company_id,worksite,created_by])
            #         for result in cursor.stored_results():
            #                 r = list(result.fetchall())
            #     type='urm'
            #     if r[0][0] == "success":
            #         messages.success(request, 'Data updated successfully !')
                
            # else : messages.error(request, 'Oops...! Something went wrong!')
                try:
                    created_by = request.session.get('user_id', '')
                    ur = request.POST.get('ur', '')
                    selected_worksite = request.POST.getlist('worksite', [])
                    company_worksite_map = {}

                    # Validate input
                    if not selected_worksite:
                        messages.error(request, 'Worksite data is missing!')
                        return redirect(f'/masters?entity={entity}&type=urm')

                    if type not in ['acu', 'acr'] or not ur:
                        messages.error(request, 'Invalid data received.')
                        return redirect(f'/masters?entity={entity}&type=urm')

                    # Parse selected worksites into company-worksite pairs
                    selected_worksite_pairs = [
                        tuple(ws.split(" - ", 1)) for ws in selected_worksite if " - " in ws
                    ]
                    if not selected_worksite_pairs:
                        messages.error(request, 'Invalid worksite format. Expected "Company Name - Worksite Name".')
                        return redirect(f'/masters?entity={entity}&type=urm')

                    valid_combinations = []
                    for company_name, worksite_name in selected_worksite_pairs:
                        try:
                            # Fetch company_id using ORM
                            company = com.objects.get(company_name=company_name)
                            company_id = company.company_id
                        except com.DoesNotExist:
                            messages.error(request, f'Company "{company_name}" does not exist.')
                            continue

                        # Check if the worksite exists for the company in SiteMaster
                        if sit.objects.filter(company_id=company_id, site_name=worksite_name).exists():
                            valid_combinations.append((company_id, worksite_name))
                        else:
                            messages.error(request, f'Worksite "{worksite_name}" does not exist for company "{company_name}".')

                    if not valid_combinations:
                        messages.error(request, 'No valid company-worksite combinations found.')

                    # Remove existing mappings
                    cursor.callproc("stp_delete_access_control", [type, ur])

                    # Insert new mappings
                    insertion_status = "failure"
                    for company_id, worksite_name in valid_combinations:
                        cursor.callproc("stp_post_access_control", [type, ur, company_id, worksite_name, created_by])
                        for result in cursor.stored_results():
                            r = list(result.fetchall())
                            if r and r[0][0] == "success":
                                insertion_status = "success"

                    # Redirect based on insertion status
                    if insertion_status == "success":
                        messages.success(request, 'Data updated successfully!')
                    else:
                        messages.error(request, 'Oops...! Something went wrong!')
                except Exception as e:
                    tb = traceback.extract_tb(e.__traceback__)
                    fun = tb[0].name
                    cursor.callproc("stp_error_log",[fun,str(e),user])  
                    messages.error(request, 'Oops...! Something went wrong!')

                    
            else : messages.error(request, 'Oops...! Something went wrong!')
                             
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method=="GET":
            return render(request,'Master/index.html', {'entity':entity,'type':type,'name':name,'header':header,'company_names':company_names,'site_name':site_name,'designation_name':designation_name,'data':data,'pre_url':pre_url})
        elif request.method=="POST":
            if entity == 'urm':
                new_url = f'/masters?entity={entity}&type=urm'
                return redirect(new_url) 
            else:
                new_url = f'/masters?entity={entity}&type={type}'
                return redirect(new_url) 
        
def gen_roster_xlsx_col(columns,month_input):
    year, month = map(int, month_input.split('-'))
    _, num_days = calendar.monthrange(year, month)
    date_columns = [(datetime(year, month, day)).strftime('%d-%m-%Y') for day in range(1, num_days + 1)]
    columns.extend(date_columns)
    return columns

def sample_xlsx(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    pre_url = request.META.get('HTTP_REFERER')
    response = ''
    global user
    user = request.session.get('user_id', '')

    try:
        # Create a new workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Sample Format'
        columns = []

        # Fetching entity and type based on request method
        if request.method == "GET":
            entity = request.GET.get('entity', '')  # Ensure entity is fetched from GET
            type = request.GET.get('type', '')
        elif request.method == "POST":
            entity = request.POST.get('entity', '')  # Ensure entity is fetched from POST
            type = request.POST.get('type', '')

        # Map the entity to a file name
        file_name = {
            'em': 'Employee Master',
            'sm': 'Worksite Master',
            'cm': 'Company Master',
            'r': 'Roster',
            'um': 'Employee Master Upload'
        }.get(entity, 'Unknown File')

        # Call stored procedure with provided parameters
        cursor.callproc("stp_get_masters", [entity, type, 'sample_xlsx', user])

        for result in cursor.stored_results():
            columns = [col[0] for col in result.fetchall()]

        # Additional logic for roster entity
        if entity == "r":
            month = request.POST.get('month', '')
            columns = gen_roster_xlsx_col(columns, month)

        # Define black border style
        black_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        # Write headers to the Excel file
        for col_num, header in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = black_border

        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        # Create HTTP response with Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="{} {}.xlsx"'.format(file_name, datetime.now().strftime("%d-%m-%Y"))
        workbook.save(response)

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])
        messages.error(request, 'Oops...! Something went wrong!')

    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()

    return response
        
# def sample_xlsx(request):
#     Db.closeConnection()
#     m = Db.get_connection()
#     cursor=m.cursor()
#     pre_url = request.META.get('HTTP_REFERER')
#     response =''
#     global user
#     user  = request.session.get('user_id', '')
#     try:
        
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.title = 'Sample Format'
#         columns = []
#         if request.method=="GET":
#             entity = request.GET.get('entity', '')
#             type = request.GET.get('type', '')
#         if request.method=="POST":
#             entity = request.POST.get('entity', '')
#             type = request.POST.get('type', '')
#         file_name = {'em': 'Employee Master','sm': 'Worksite Master','cm': 'Company Master','r': 'Roster','um': 'Employee Master Upload'}[entity]
#         cursor.callproc("stp_get_masters", [entity, type, 'sample_xlsx',user])
#         for result in cursor.stored_results():
#             columns = [col[0] for col in result.fetchall()]
#         # columns = ['Column 1', 'Column 2', 'Column 3']
#         if entity == "r":
#             month = request.POST.get('month', '')
#             columns = gen_roster_xlsx_col(columns,month)

#         black_border = Border(
#             left=Side(border_style="thin", color="000000"),
#             right=Side(border_style="thin", color="000000"),
#             top=Side(border_style="thin", color="000000"),
#             bottom=Side(border_style="thin", color="000000")
#         )
        
#         for col_num, header in enumerate(columns, 1):
#             cell = sheet.cell(row=1, column=col_num)
#             cell.value = header
#             cell.font = Font(bold=True)
#             cell.border = black_border
        
#         for col in sheet.columns:
#             max_length = 0
#             column = col[0].column_letter  
#             for cell in col:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
                    
#             adjusted_width = max_length + 2 
#             sheet.column_dimensions[column].width = adjusted_width  
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename="' + str(file_name) +" "+str(datetime.now().strftime("%d-%m-%Y")) + '.xlsx"'
#         workbook.save(response)
    
#     except Exception as e:
#         tb = traceback.extract_tb(e.__traceback__)
#         fun = tb[0].name
#         cursor.callproc("stp_error_log",[fun,str(e),user])  
#         messages.error(request, 'Oops...! Something went wrong!')
#     finally:
#         cursor.close()
#         m.commit()
#         m.close()
#         Db.closeConnection()
#         return response      

@login_required  
def roster_upload(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    global user
    user  = request.session.get('user_id', '')
    if request.method == 'POST' and request.FILES.get('roster_file'):
        try:
            excel_file = request.FILES['roster_file']
            file_name = excel_file.name
            df = pd.read_excel(excel_file)

            entity = request.POST.get('entity', '')
            type = request.POST.get('type', '')
            company_id = request.POST.get('company_id', '')
            month_input  =str(request.POST.get('month_year', ''))
            total_rows = len(df)
            update_count = error_count = success_count = 0
            checksum_id = None
            worksites = []

            if entity == 'r':
                year, month = map(int, month_input.split('-'))
                _, num_days = calendar.monthrange(year, month)
                date_columns = [(datetime(year, month, day)).strftime('%d-%m-%Y') for day in range(1, num_days + 1)]
                cursor.callproc("stp_get_masters", [entity, type, 'sample_xlsx',user])
                for result in cursor.stored_results():
                    start_columns = [col[0] for col in result.fetchall()]

                if not all(col in df.columns for col in start_columns + date_columns):
                    messages.error(request, 'Oops...! The uploaded Excel file does not contain the required columns.!')
                    return redirect(f'/masters?entity={entity}&type={type}')
                
                cursor.callproc('stp_insert_checksum', ('roster',company_id,month,year,file_name))
                for result in cursor.stored_results():
                    c = list(result.fetchall())
                checksum_id = c[0][0]
                
                for index,row in df.iterrows():
                    employee_id = row.get('Employee Id', '')
                    employee_name = row.get('Employee Name', '')
                    worksite  = row.get('Worksite', '')
                    
                    for date_col in date_columns:
                        shift_date = datetime.strptime(date_col, '%d-%m-%Y').date()
                        shift_time = row.get(date_col) 
                        if pd.isna(shift_time):
                            shift_time = None
                        params = (str(employee_id),employee_name,int(company_id),worksite,shift_date,shift_time,checksum_id,user)
                        cursor.callproc('stp_insert_roster', params)
                        for result in cursor.stored_results():
                            r = list(result.fetchall())
                        if r[0][0] not in ("success", "updated"):
                            if worksite not in worksites:
                                worksites.append(worksite)
                            error_message = str(r[0][0])
                            error_params = ('roster', company_id,worksite,file_name,shift_date,error_message,checksum_id)
                            cursor.callproc('stp_insert_error_log', error_params)
                            messages.error(request, "Errors occurred during upload. Please check error logs.")
                    if r[0][0] == "success": success_count += 1
                    elif r[0][0] == "updated": update_count += 1  
                    else: error_count += 1
                checksum_msg = f"Total Rows Processed: {total_rows}, Successful Entries: {success_count}, Updates: {update_count}, Errors: {error_count}"
                cursor.callproc('stp_update_checksum', ('roster',company_id,', '.join(worksites),month,year,file_name,checksum_msg,error_count,update_count,checksum_id))
                if error_count == 0 and update_count == 0 and success_count > 0:
                    messages.success(request, f"All data uploaded successfully!.")
                elif error_count == 0 and success_count == 0 and update_count > 0:
                    messages.warning(request, f"All data updated successfully!.")
                else:
                    messages.warning(request, f"The upload processed {total_rows} rows, resulting in {success_count} successful entries, {update_count} updates, and {error_count} errors; please check the error logs for details.")
                    
        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name
            cursor.callproc("stp_error_log", [fun, str(e), user])  
            messages.error(request, 'Oops...! Something went wrong!')
            m.commit()   

        finally:
            cursor.close()
            m.close()
            Db.closeConnection()
            return redirect(f'/masters?entity={entity}&type={type}')     
        
@login_required        
def site_master(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    global user
    user  = request.session.get('user_id', '')
    try:
        
        if request.method == "GET":
            # cursor.callproc("stp_get_roster_type")
            # for result in cursor.stored_results():
            #     roster_types = list(result.fetchall())
                # Call stored procedure to get company names
            cursor.callproc("stp_get_userwise_dropdown", [user,'company'])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())
            cursor.callproc("stp_get_state_names")
            for result in cursor.stored_results():
                state_names = list(result.fetchall())
            cursor.callproc("stp_get_city_names")
            for result in cursor.stored_results():
                city_names = list(result.fetchall())
                
            site_id = request.GET.get('site_id', '')
            # site_id = decrypt_parameter(str(site_id))
            if site_id == "0":
                if request.method == "GET":
                    context = {'company_names': company_names, 'state_names': state_names,'site_id':site_id}

            else:
                # site_id1 = request.GET.get('site_id', '')
                # site_id = decrypt_parameter(site_id1)

                # cursor.callproc("stp_edit_site_master",(site_id, )) 
                # for result in cursor.stored_results():
                #     data = result.fetchall()[0]
                site_id1 = request.GET.get('site_id', '')
                site_id = decrypt_parameter(site_id1)
                cursor.callproc("stp_edit_site_master", (site_id,)) 
                for result in cursor.stored_results():
                    data = result.fetchall()[0] 

                    context = {
                        'company_names': company_names,
                        'state_names': state_names,
                        'city_names': city_names,
                        'site_id': site_id[0],
                        'site_name': data[1],
                        'site_address': data[2],
                        'pincode': data[3],
                        'state_id': data[4],
                        'contact_person_name': data[5],
                        'contact_person_email': data[6],
                        'contact_person_mobile_no': data[7],
                        'is_active': data[8],
                        'company_name': data[9],
                        'city': data[10],
                    }
                
            
        if request.method == "POST":
            siteId = request.POST.get('site_id', '')
            if siteId == "0":
                # response_data = {"status": "fail"}
                try:
                    siteName = request.POST.get('siteName', '')
                    siteAddress = request.POST.get('siteAddress', '')
                    pincode = request.POST.get('pincode', '')
                    contactPersonName = request.POST.get('contactPersonName', '')
                    contactPersonEmail = request.POST.get('contactPersonEmail', '')
                    contactPersonMobileNo = request.POST.get('Number', '')  
                    # is_active = request.POST.get('status_value', '') 
                    # noOfDays = request.POST.get('FieldDays', '')  
                    # notificationTime = request.POST.get('notificationTime', '')
                    # ReminderTime = request.POST.get('ReminderTime', '')
                    companyId = request.POST.get('company_id', '')  
                    stateId = request.POST.get('state_id', '')  
                    cityId = request.POST.get('city_id', '')  
                    # rosterType = request.POST.get('roster_type', '')
                
                    params = [
                        siteName, 
                        siteAddress, 
                        pincode, 
                        contactPersonName, 
                        contactPersonEmail, 
                        contactPersonMobileNo, 
                        # is_active,
                        # noOfDays, 
                        # notificationTime, 
                        # ReminderTime,
                        stateId,
                        cityId,
                        companyId,
                        user
                        # rosterType
                    ]
                    
                    cursor.callproc("stp_insert_site_master", params)
                    for result in cursor.stored_results():
                            datalist = list(result.fetchall())
                    if datalist[0][0] == "success":
                        messages.success(request, 'Data successfully entered !')
                    else: messages.error(request, datalist[0][0])
                except Exception as e:
                    tb = traceback.extract_tb(e.__traceback__)
                    fun = tb[0].name
                    cursor.callproc("stp_error_log", [fun, str(e), user])  
                    messages.error(request, 'Oops...! Something went wrong!')
            else:
                if request.method == "POST" :
                    siteId = request.POST.get('site_id', '')
                    siteName = request.POST.get('siteName', '')
                    siteAddress = request.POST.get('siteAddress', '')
                    pincode = request.POST.get('pincode', '')
                    contactPersonName = request.POST.get('contactPersonName', '')
                    contactPersonEmail = request.POST.get('contactPersonEmail', '')
                    contactPersonMobileNo = request.POST.get('Number', '')  
                    # noOfDays = request.POST.get('FieldDays', '') 
                    isActive = request.POST.get('status_value', '')
                    # notificationTime = request.POST.get('notificationTime', '')
                    # ReminderTime = request.POST.get('ReminderTime', '')
                    companyId = request.POST.get('company_id', '')  
                    stateId = request.POST.get('state_id', '')  
                    cityId = request.POST.get('city_id', '')  
                    # Rostertype = request.POST.get('roster_type', '')
                    
                    params = [siteId,siteName,siteAddress,pincode,contactPersonName,contactPersonEmail,
                                        contactPersonMobileNo,isActive,companyId,stateId,cityId,user]
                    cursor.callproc("stp_update_site_master",params) 
                    messages.success(request, "Data updated successfully...!")

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
            
        if request.method=="GET":
            return render(request, "Master/master/site_master.html", context)
        elif request.method=="POST":  
            return redirect( f'/masters?entity=sm&type=i')
        
@login_required      
def company_master(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    global user
    user  = request.session.get('user_id', '')
    try:
        
        if request.method == "GET":
        
            company_id = request.GET.get('company_id', '')
            if company_id == "0":
                if request.method == "GET":
                    context = {'company_id':company_id}
            else:
                company_id1 = request.GET.get('company_id', '')
                company_id= decrypt_parameter(company_id1)
                cursor.callproc("stp_edit_company_master", (company_id,))  # Note the comma to make it a tuple
                for result in cursor.stored_results():
                    data = result.fetchall()[0]  
                        
                    context = {
                        'company_id' : data[0],
                        'company_name': data[1],
                        'company_address': data[2],
                        'pincode': data[3],
                        'contact_person_name': data[4],
                        'contact_person_email': data[5], 
                        'contact_person_mobile_no': data[6],
                        'is_active':data[7]
                    }

        if request.method == "POST" :
            company_id = request.POST.get('company_id', '')
            if company_id == '0':
                response_data = {"status": "fail"}
                company_name = request.POST.get('company_name', '')
                company_address = request.POST.get('company_address', '')
                pincode = request.POST.get('pincode', '')
                contact_person_name = request.POST.get('contact_person_name', '')
                contact_person_email = request.POST.get('contact_person_email', '')
                contact_person_mobile_no = request.POST.get('contact_person_mobile_no', '') 
                # is_active = request.POST.get('status_value', '') 
                params = [
                    company_name, 
                    company_address, 
                    pincode, 
                    contact_person_name,
                    contact_person_email,
                    contact_person_mobile_no
                    # is_active
                ]
                cursor.callproc("stp_insert_company_master", params)
                for result in cursor.stored_results():
                        datalist = list(result.fetchall())
                if datalist[0][0] == "success":
                    messages.success(request, 'Data successfully entered !')
                else: messages.error(request, datalist[0][0])
            else :
                company_id = request.POST.get('company_id', '')
                company_name = request.POST.get('company_name', '')
                company_address = request.POST.get('company_address', '')
                pincode = request.POST.get('pincode', '')
                contact_person_name = request.POST.get('contact_person_name', '')
                contact_person_email = request.POST.get('contact_person_email', '')
                contact_person_mobile_no = request.POST.get('contact_person_mobile_no', '') 
                is_active = request.POST.get('status_value', '') 
                   
                params = [company_id,company_name,company_address,pincode,contact_person_name,contact_person_email,
                                            contact_person_mobile_no,is_active]    
                cursor.callproc("stp_update_company_master",params) 
                messages.success(request, "Data updated successfully ...!")
                
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()

        encrypted_id = encrypt_parameter(company_id)
            
        if request.method=="GET":
            return render(request, "Master/master/company_master.html", context)
        elif request.method == "POST":
            return redirect(f'/masters?entity=cm&type=i')

@login_required        
def employee_master(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    global user
    
    user = request.session.get('user_id', '')
    try:
        
        if request.method == "GET":
            id = request.GET.get('id', '')
            

            cursor.callproc("stp_get_employee_status")
            for result in cursor.stored_results():
                employee_status = list(result.fetchall())
            if id != '0':  
                id1 = decrypt_parameter(id)
                cursor.callproc("sto_get_employee_site", [id1,])    
                for result in cursor.stored_results():
                    site_name = list(result.fetchall())
            else:
                site_name = [] 
            cursor.callproc("stp_get_company_site_name",[user])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())
            cursor.callproc("stp_get_dropdown_values",('states',))
            for result in cursor.stored_results():
                state_names = list(result.fetchall())
            cursor.callproc("stp_get_dropdown_values",('gender',))
            for result in cursor.stored_results():
                gender = list(result.fetchall())
            cursor.callproc("stp_get_dropdown_values",('designation',))
            for result in cursor.stored_results():
                designation_name = list(result.fetchall())
            if id == "0":
                if request.method == "GET":
                    context = {'id':id, 'employee_status':employee_status, 'employee_status_id': '','gender':gender ,'company_names':company_names,'state_names':state_names,'designation_name':designation_name}

            else:
                id1 = request.GET.get('id', '')
                id = decrypt_parameter(id1)
                cursor.callproc("stp_edit_employee_master", (id,))
                for result in cursor.stored_results():
                    data = result.fetchall()[0]  
                    context = {
                        'id':id, 
                        'employee_status':employee_status, 
                        'site_name':site_name,
                        'designation_name':designation_name,
                        'gender':gender ,
                        'company_names':company_names,
                        'state_names':state_names,
                        'employee_status':employee_status,
                        'employee_id' : data[0],
                        'employee_name': data[1],
                        'mobile_no': data[2],
                        'email':data[3], 
                        'gender_value': data[4],
                        'state_id_value':data[5],
                        'city':data[6],
                        'address':data[7],
                        'pincode':data[8],
                        'account_holder_name':data[9],
                        'account_no':data[10],
                        'bank_name':data[11],
                        'branch_name':data[12],
                        'ifsc_code':data[13],
                        'pf_no':data[14],
                        'uan_no':data[15],
                        'esic':data[16],
                        'employment_status_id': data[17],
                        'handicapped_value': data[18],
                        'is_active': data[19],     
                        'company_id_value':data[20],
                    }
                    employee_id = context['employee_id']
                    designation_list = employee_designation.objects.filter(employee_id=employee_id)
                    context['selected_designation'] = json.dumps(list(map(str, designation_list.values_list('designation_id', flat=True)))) 


                    site_list = employee_site.objects.filter(employee_id=employee_id)
                    context['selected_site'] = json.dumps(list(map(str,site_list.values_list('site_id', flat=True)))) 



        if request.method == "POST" :
            id = request.POST.get('id', '')
            if id == '0':

                employeeId = request.POST.get('employee_id', '')
                employeeName = request.POST.get('employee_name', '')
                mobile_no = request.POST.get('mobile_no', '')
                email= request.POST.get('email', '')
                gender = request.POST.get('gender', '')
                handicapped = request.POST.get('handicapped_value', '')
                address = request.POST.get('address', '')
                city = request.POST.get('city', '')
                state1 = request.POST.get('state_id', '')
                pincode = request.POST.get('pincode', '')
                company_id1 = request.POST.get('company_id', '')
                account_holder_name = request.POST.get('account_holder_name', '')
                account_no = request.POST.get('account_no', '')
                bank_name = request.POST.get('bank_name', '')
                branch_name = request.POST.get('branch_name', '')
                ifsc_code = request.POST.get('ifsc_code', '')
                pf_no = request.POST.get('pf_no', '')
                uan_no = request.POST.get('uan_no', '')
                esic = request.POST.get('esic', '')
                # employeeStatus = request.POST.get('employee_status_id', '')
                # activebtn = request.POST.get('status_value', '')
                designation_list = request.POST.getlist('designation_name', '')
                site_list = request.POST.getlist('site_name', '')
                for designation in designation_list:
                    designation_id = designation_master.objects.get(designation_id=int(designation))
                    employee_designation.objects.create(employee_id=employeeId,designation_id=designation_id)
                for site in site_list:
                    site_id = sit.objects.get(site_id=int(site))
                    employee_site.objects.create(employee_id=employeeId,site_id=site_id)

                params = [
                    employeeId, 
                    employeeName, 
                    mobile_no,
                    email, 
                    address,
                    city,
                    state1,
                    pincode,
                    company_id1,
                    account_holder_name,
                    account_no,
                    bank_name,
                    branch_name,
                    ifsc_code,
                    pf_no,
                    uan_no,
                    esic,
                    gender,
                    handicapped,
                    user
                    # employeeStatus,
                    # activebtn
                ]
                
                cursor.callproc("stp_insert_employee_master", params)
                for result in cursor.stored_results():
                        datalist = list(result.fetchall())
                if datalist[0][0] == "success":
                    messages.success(request, 'Data successfully Saved !')
                else: messages.error(request, datalist[0][0])
            else:
                id = request.POST.get('id', '')
                employee_id = request.POST.get('employee_id', '')
                employee_name = request.POST.get('employee_name', '')
                mobile_no = request.POST.get('mobile_no', '')
                email = request.POST.get('email', '')
                gender = request.POST.get('gender', '')
                handicapped = request.POST.get('handicapped_value', '')
                address = request.POST.get('address', '')
                city = request.POST.get('city', '')
                state1 = request.POST.get('state_id', '')
                pincode = request.POST.get('pincode', '')
                company_id1 = request.POST.get('company_id', '')
                account_holder_name = request.POST.get('account_holder_name', '')
                account_no = request.POST.get('account_no', '')
                bank_name = request.POST.get('bank_name', '')
                branch_name = request.POST.get('branch_name', '')
                ifsc_code = request.POST.get('ifsc_code', '')
                pf_no = request.POST.get('pf_no', '')
                uan_no = request.POST.get('uan_no', '')
                esic = request.POST.get('esic', '')
                employee_status = request.POST.get('employment_status', '')
                is_active = 1 if request.POST.get('is_active', '') == 'on' else 0 
                designation_list = request.POST.getlist('designation_name', '')
                site_list = request.POST.getlist('site_name', '')
                            
                params = [ id,employee_id, employee_name, mobile_no,email, address,city,state1,pincode,account_holder_name,account_no,bank_name,branch_name,ifsc_code,pf_no,uan_no,esic,gender,handicapped,company_id1,employee_status,is_active,user]    
                # cursor.callproc("stp_update_employee_master",params) 
                # messages.success(request, "Data successfully Updated!")

                 
                cursor.callproc("stp_update_employee_master", params)
                for result in cursor.stored_results():
                        datalist = list(result.fetchall())
               # Updating designations for the employee
                        
                employee_designation.objects.filter(employee_id=employee_id).delete()  # Replace with your actual model

                # Insert new rows with the employee_id and designation_id
                for designation_idd in designation_list:
                    # Create a new instance of the model with the employee_id and designation_id
                    designation_id1 = get_object_or_404(designation_master, designation_id=designation_idd)
                    new_entry = employee_designation(employee_id=employee_id, designation_id=designation_id1)
                    new_entry.save()

                # Deleting the existing sites for the employee
                employee_site.objects.filter(employee_id=employee_id).delete()

                # Updating sites for the employee
                for site_idd in site_list:
                    # Fetching the correct site using the 'id' field (or the correct field name)
                    site_id1 = get_object_or_404(sit, site_id=site_idd)
                    
                    # Creating a new employee-site relation
                    new_entry = employee_site(employee_id=employee_id, site_id=site_id1)
                    new_entry.save()

                
                if datalist[0][0] == "success":
                    messages.success(request, 'Data successfully Updated !')
                else: messages.error(request, datalist[0][0])

    except Exception as e:
        print(f"Error: {e}")  # Temporary for debugging
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])  
        messages.error(request, 'Oops...! Something went wrong!')


    
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method=="GET":
            return render(request, "Master/master/employee_master.html", context)
        elif request.method=="POST":  
            return redirect(f'/masters?entity=em&type=i')

@login_required  
def upload_excel(request):

    if request.method == 'POST' and request.FILES.get('excelFile'):
        excel_file = request.FILES['excelFile']
        file_name = excel_file.name
        df = pd.read_excel(excel_file)
        total_rows = len(df)
        update_count = error_count = success_count = 0
        checksum_id = None
        r=None
        global user
        user  = request.session.get('user_id', '')
        try:
            Db.closeConnection()
            m = Db.get_connection()
            cursor = m.cursor()
            entity = request.POST.get('entity', '')
            type = request.POST.get('type', '')
            company_id1 = request.POST.get('company_id', None)
            cursor.callproc("stp_get_masters", [entity, type, 'sample_xlsx',user])
            for result in cursor.stored_results():
                columns = [col[0] for col in result.fetchall()]
            if not all(col in df.columns for col in columns):
                messages.error(request, 'Oops...! The uploaded Excel file does not contain the required columns.!')
                return redirect(f'/masters?entity={entity}&type={type}')
            upload_for = {'em': 'employee master','sm': 'site master','cm': 'company master','r': 'roster'}[entity]
            cursor.callproc('stp_insert_checksum', (upload_for,company_id1,str(datetime.now().month),str(datetime.now().year),file_name))
            for result in cursor.stored_results():
                c = list(result.fetchall())
            checksum_id = c[0][0]

            if entity == 'em':
                success_count = 0
                error_count = 0
                update_count = 0
                for index, row in df.iterrows():
                    row_error_found = False
                    row_was_updated = False
                    # Retrieve employee_id from the current row
                    employee_id1 = row['Employee Id']  # Ensure this matches your DataFrame's column name

                    row_filtered = row.drop(['worksite', 'designation'], errors='ignore')
                    params = tuple(str(row_filtered.get(column, '')) for column in columns if column not in ['worksite', 'designation'])
                    
                    # After validation, add company_id to params
                    params += (str(company_id1),)
                    params += (str(user),)
                    merged_list = list(zip(columns, params))

                    print(merged_list)

                    # Loop through each (column, value) pair in the merged_list for custom validation
                    for column, value in merged_list:
                        # Convert value to string and lowercase, then print for debugging
                        if isinstance(value, str):
                            value = value.strip().lower()

                        # Skip validation and error logging for 'Designation' and 'Worksite' columns
                        if column in ['Designation', 'Worksite']:
                            continue  # Skip to the next iteration if the column is 'designation' or 'worksite'

                        # Call the stored procedure with employee_id1
                        cursor.callproc('stp_employee_validation', [column, value, employee_id1])
                        for result in cursor.stored_results():
                            r = list(result.fetchall())
                            if r and r[0][0] not in ("", None, " ", "Success"):
                                error_message = str(r[0][0])
                                # Ensure proper logging for errors
                                cursor.callproc('stp_insert_error_log', [upload_for, company_id1, file_name, datetime.now().date(), error_message, checksum_id, employee_id1])
                                error_count += 1
                                row_error_found = True 

                    if not row_error_found:

                        employee_exists = sc_employee_master.objects.filter(employee_id=employee_id1).exists()
                        # Check if the row is an update or a new insert

                        if employee_exists:
                            cursor.callproc('stp_update_employee_master_excel', params)
                            for result in cursor.stored_results():
                                update_result = result.fetchone()
                        else:
                            cursor.callproc('stp_employeeinsertexcel', params)
                            for result in cursor.stored_results():
                                update_result = result.fetchone()

                        
                            if update_result == "Updated":
                                update_count += 1  
                                row_was_updated = True
                            elif update_result == "Success":
                                success_count += 1  

                
                df.columns = df.columns.str.strip()  


                if 'Employee Id' in df.columns and 'Designation' in df.columns:
                    # print("Both 'employee id' and 'designation' found in DataFrame.")

                   
                    df_designations = df[['Employee Id', 'Designation']].dropna(subset=['Designation']).copy()
                    df_designations['company_id'] = company_id1  
                    
                    for index, row in df_designations.iterrows():
                        employee_id = row['Employee Id']
                        designations = row['Designation'].split(',')  

                       
                        for designation in designations:
                            designation = designation.strip()  
                            
                            
                            cursor.callproc('stp_insert_employee_designation', [employee_id, designation, company_id1])

                          
                            for result in cursor.stored_results():
                                r = list(result.fetchall())

                            if r:
                                result_value = r[0][0]  

                                if result_value == "success":
                                    success_count += 1
                                elif result_value == "updated":
                                    update_count += 1
                                elif result_value.startswith("error"):
                                    # Log the error message in your error log table
                                    error_message = result_value  # The error message from the procedure
                                    cursor.callproc('stp_insert_error_log', [
                                        upload_for, company_id1, file_name, datetime.now().date(), error_message, checksum_id,employee_id
                                    ])
                                    error_count += 1  # Increment error count for errors

                else:
                    # In case 'employee id' or 'designation' is missing in df.columns
                    print("Required columns 'employee id' or 'designation' not found in the DataFrame.")

                       #   this is for worksite insert  
                if 'Employee Id' in df.columns and 'Worksite' in df.columns:
                    # print("Both 'employee id' and 'worksite' found in DataFrame.")

                    # Create a second DataFrame for 'employee_id', 'designation', and company_id
                    df_worksite = df[['Employee Id', 'Worksite']].dropna(subset=['Worksite']).copy()
                    df_worksite['company_id'] = company_id1  # Add company_id to df_designations

                    # Loop through each row in the DataFrame
                    for index, row in df_worksite.iterrows():
                        employee_id = row['Employee Id']
                        worksite = row['Worksite'].split(',')  # Assuming multiple designations are comma-separated

                        # Loop through each designation and insert it into the employee_designation table
                        for worksite in worksite:
                            worksite = worksite.strip()  # Trim spaces from the designation
                            
                            # Call the stored procedure to insert the designation for the employee
                            cursor.callproc('stp_insert_employee_site', [employee_id, worksite, company_id1])

                            # Fetch all stored results to check the response from the procedure
                            for result in cursor.stored_results():
                                r = list(result.fetchall())

                            if r:
                                result_value = r[0][0]  # Fetch the result from the stored procedure

                                if result_value == "success":
                                    success_count += 1
                                elif result_value == "updated":
                                    update_count += 1
                                elif result_value.startswith("error"):
                                    # Log the error message in your error log table
                                    error_message = result_value  # The error message from the procedure
                                    cursor.callproc('stp_insert_error_log', [
                                        upload_for, company_id1, file_name, datetime.now().date(), error_message, checksum_id,employee_id1
                                    ])
                                    error_count += 1  # Increment error count for errors

                else:
                    # In case 'employee id' or 'designation' is missing in df.columns
                    print("Required columns 'employee id' or 'worksite' not found in the DataFrame.")
            elif entity == 'sm':
                for index, row in df.iterrows():
                    state_name = str(row.get('State', ''))
                    city_name = str(row.get('City', ''))
                    
                    try:
                        state_obj = StateMaster.objects.get(state_name=state_name)
                        state_id = state_obj.state_id
                    except StateMaster.DoesNotExist:
                        state_id = None  # State not found
                    
                    try:
                        city_obj = CityMaster.objects.get(city_name=city_name)
                        city_id = city_obj.city_id
                    except CityMaster.DoesNotExist:
                        city_id = None  # City not found

                    # Check if state_id or city_id is None and log error
                    if state_id is None:
                        error_message = f"Please provide a Valid State for the Worksite '{row.get('Site Name')}' at row number {index + 1}"
                        cursor.callproc('stp_site_error_log', [upload_for, company_id1, file_name, datetime.now().date(), error_message, checksum_id, ''])
                    
                    if city_id is None:
                        error_message = f"Please provide a Valid City for the Worksite '{row.get('Site Name')}' at row number {index + 1}"
                        cursor.callproc('stp_site_error_log', [upload_for, company_id1, file_name, datetime.now().date(), error_message, checksum_id, ''])

                    # Only proceed to insert into site_master if state_id and city_id are not None
                    if state_id is not None and city_id is not None:
                        # Get the other parameters
                        params = tuple(str(row.get(column, '')) for column in columns if column not in ['State', 'City'])
                        params += (str(state_id), str(city_id))
                        params += (str(company_id1),)
                        params += (str(user),)

                        # Insert data into the site_master table
                        cursor.callproc('stp_insert_site_master_excel', params)
                        
                        # Check the result from the stored procedure
                        for result in cursor.stored_results():
                            r = list(result.fetchall())

                        if r[0][0] not in ("success", "updated"):
                            # Log error if the result is not success or updated
                            cursor.callproc('stp_site_error_log', [upload_for, company_id1, file_name, datetime.now().date(), str(r[0][0]), checksum_id, str(row.get('Site Name'))])

                    
                    # Increment counters based on the result
                    if r[0][0] == "success":
                        success_count += 1
                    elif r[0][0] == "updated":
                        update_count += 1
                    else:
                        error_count += 1
            elif entity == 'cm':
                for index,row in df.iterrows():
                    params = tuple(str(row.get(column, '')) for column in columns)
                    cursor.callproc('stp_insert_company_master', params)
                    for result in cursor.stored_results():
                            r = list(result.fetchall())
                    if r[0][0] not in ("success", "updated"):
                        cursor.callproc('stp_insert_error_log', [upload_for, company_id1,'',file_name,datetime.now().date(),str(r[0][0]),checksum_id])
                    if r[0][0] == "success": success_count += 1 
                    elif r[0][0] == "updated": update_count += 1  
                    else: error_count += 1
            checksum_msg = f"Total Rows Processed: {total_rows}, Successful Entries: {success_count}" f"{f', Updates: {update_count}' if update_count > 0 else ''}" f"{f', Errors: {error_count}' if error_count > 0 else ''}"
            cursor.callproc('stp_update_checksum', (upload_for,company_id1,'',str(datetime.now().month),str(datetime.now().year),file_name,checksum_msg,error_count,update_count,checksum_id,''))
            if error_count == 0 and update_count == 0 and success_count > 0:
                messages.success(request, f"All data uploaded successfully!.")
            elif error_count == 0 and success_count == 0 and update_count > 0:
                messages.success(request, f"All data updated successfully!.")
            else:messages.warning(request, f"The upload processed {total_rows} rows, resulting in {success_count} successful entries"  f"{f', {update_count} updates' if update_count > 0 else ''}" f", and {error_count} errors; please check the error logs for details.")
                   
        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name
            cursor.callproc("stp_error_log", [fun, str(e), user])  
            messages.error(request, 'Oops...! Something went wrong!')
            m.commit()   
        finally:
            cursor.close()
            m.close()
            Db.closeConnection()
            return redirect(f'/masters?entity={entity}&type=i')

# @login_required  
# def upload_excel(request):

#      if request.method == 'POST' and request.FILES.get('excelFile'):
#         excel_file = request.FILES['excelFile']
#         file_name = excel_file.name
#        # Read the Excel file into three separate DataFrames
#         df = pd.read_excel(excel_file)
        
#         # Concatenate the three DataFrames into one
      
#         # Calculate the total number of rows
#         total_rows = len(df)
#         total_columns = len(df.columns)

#         update_count = error_count = success_count = 0
#         checksum_id = None
#         r=None
#         global user
#         user  = request.session.get('user_id', '')
#         try:
#             Db.closeConnection()
#             m = Db.get_connection()
#             cursor = m.cursor()
#             entity = request.POST.get('entity', '')
#             type = request.POST.get('type', '')
#             company_id = request.POST.get('company_id','')
#             # state_id=request.POST.get('state_id', None)
#             cursor.callproc("stp_get_masters", [entity, type, 'sample_xlsx',user])
#             for result in cursor.stored_results():
#                 columns = [col[0] for col in result.fetchall()]
#             # if not all(col in df.columns for col in columns):
#             #     messages.error(request, 'Oops...! The uploaded Excel file does not contain the required columns.!')
#             #     return redirect(f'/masters?entity={entity}&type={type}')
#             upload_for = {'em': 'employee master','um': 'employee master upload','sm': 'site master','cm': 'company master','r': 'roster'}[entity]
#             cursor.callproc('stp_insert_checksum', (upload_for,company_id,str(datetime.now().month),str(datetime.now().year),file_name))
#             for result in cursor.stored_results():
#                 c = list(result.fetchall())
#             checksum_id = c[0][0]
        
#             if entity == 'um':
#                 success_count = 0
#                 error_count = 0
#                 update_count = 0
#                 for index, row in df.iterrows():
#                     row_error_found = False
#                     row_was_updated = False
#                     # Retrieve employee_id from the current row
#                     employee_id1 = row['Employee Id']  # Ensure this matches your DataFrame's column name

#                     row_filtered = row.drop(['worksite', 'designation'], errors='ignore')
#                     params = tuple(str(row_filtered.get(column, '')) for column in columns if column not in ['worksite', 'designation'])
                    
#                     # After validation, add company_id to params
#                     params += (str(company_id),)
#                     merged_list = list(zip(columns, params))

#                     print(merged_list)

#                     # Loop through each (column, value) pair in the merged_list for custom validation
#                     for column, value in merged_list:
#                         # Convert value to string and lowercase, then print for debugging
#                         if isinstance(value, str):
#                             value = value.strip().lower()

#                         # Skip validation and error logging for 'Designation' and 'Worksite' columns
#                         if column in ['Designation', 'Worksite']:
#                             continue  # Skip to the next iteration if the column is 'designation' or 'worksite'

#                         # Call the stored procedure with employee_id1
#                         cursor.callproc('stp_employee_validation', [column, value, employee_id1])
#                         for result in cursor.stored_results():
#                             r = list(result.fetchall())
#                             if r and r[0][0] not in ("", None, " ", "Success"):
#                                 error_message = str(r[0][0])
#                                 # Ensure proper logging for errors
#                                 cursor.callproc('stp_insert_error_log', [upload_for, company_id, file_name, datetime.now().date(), error_message, checksum_id, employee_id1])
#                                 error_count += 1
#                                 row_error_found = True 

#                     if not row_error_found:
#                         # Check if the row is an update or a new insert
#                         cursor.callproc('stp_employeeinsertexcel', params)
#                         for result in cursor.stored_results():
#                             update_result = result.fetchone()
#                             if update_result == "Updated":
#                                 update_count += 1  # Increment update count
#                                 row_was_updated = True
#                             elif update_result == "Success":
#                                 success_count += 1  

                
#                 df.columns = df.columns.str.strip()  


#                 if 'Employee Id' in df.columns and 'Designation' in df.columns:
#                     # print("Both 'employee id' and 'designation' found in DataFrame.")

                   
#                     df_designations = df[['Employee Id', 'Designation']].dropna(subset=['Designation']).copy()
#                     df_designations['company_id'] = company_id  
                    
#                     for index, row in df_designations.iterrows():
#                         employee_id = row['Employee Id']
#                         designations = row['Designation'].split(',')  

                       
#                         for designation in designations:
#                             designation = designation.strip()  
                            
                            
#                             cursor.callproc('stp_insert_employee_designation', [employee_id, designation, company_id])

                          
#                             for result in cursor.stored_results():
#                                 r = list(result.fetchall())

#                             if r:
#                                 result_value = r[0][0]  

#                                 if result_value == "success":
#                                     success_count += 1
#                                 elif result_value == "updated":
#                                     update_count += 1
#                                 elif result_value.startswith("error"):
#                                     # Log the error message in your error log table
#                                     error_message = result_value  # The error message from the procedure
#                                     cursor.callproc('stp_insert_error_log', [
#                                         upload_for, company_id, file_name, datetime.now().date(), error_message, checksum_id,employee_id
#                                     ])
#                                     error_count += 1  # Increment error count for errors

#                 else:
#                     # In case 'employee id' or 'designation' is missing in df.columns
#                     print("Required columns 'employee id' or 'designation' not found in the DataFrame.")

#                        #   this is for worksite insert  
#                 if 'Employee Id' in df.columns and 'Worksite' in df.columns:
#                     # print("Both 'employee id' and 'worksite' found in DataFrame.")

#                     # Create a second DataFrame for 'employee_id', 'designation', and company_id
#                     df_worksite = df[['Employee Id', 'Worksite']].dropna(subset=['Worksite']).copy()
#                     df_worksite['company_id'] = company_id  # Add company_id to df_designations

#                     # Loop through each row in the DataFrame
#                     for index, row in df_worksite.iterrows():
#                         employee_id = row['Employee Id']
#                         worksite = row['Worksite'].split(',')  # Assuming multiple designations are comma-separated

#                         # Loop through each designation and insert it into the employee_designation table
#                         for worksite in worksite:
#                             worksite = worksite.strip()  # Trim spaces from the designation
                            
#                             # Call the stored procedure to insert the designation for the employee
#                             cursor.callproc('stp_insert_employee_site', [employee_id, worksite, company_id])

#                             # Fetch all stored results to check the response from the procedure
#                             for result in cursor.stored_results():
#                                 r = list(result.fetchall())

#                             if r:
#                                 result_value = r[0][0]  # Fetch the result from the stored procedure

#                                 if result_value == "success":
#                                     success_count += 1
#                                 elif result_value == "updated":
#                                     update_count += 1
#                                 elif result_value.startswith("error"):
#                                     # Log the error message in your error log table
#                                     error_message = result_value  # The error message from the procedure
#                                     cursor.callproc('stp_insert_error_log', [
#                                         upload_for, company_id, file_name, datetime.now().date(), error_message, checksum_id,employee_id1
#                                     ])
#                                     error_count += 1  # Increment error count for errors

#                 else:
#                     # In case 'employee id' or 'designation' is missing in df.columns
#                     print("Required columns 'employee id' or 'worksite' not found in the DataFrame.")


                          
#             elif entity == 'sm':
#                 for index, row in df.iterrows():
#                     row_error_found = False
#                     row_was_updated = False
#                     site_id1 = row['Site Name']
#                     params = tuple(str(row.get(column, '')) for column in columns)
#                     params += (str(company_id),)

#                     # Iterate through the columns to validate each entry
#                     for column in columns:  # Added this line to define 'column'
#                         value = row[column]  # Get the value from the current row for the current column
#                         cursor.callproc('stp_site_validation', [column, value, site_id1])
#                         for result in cursor.stored_results():
#                             r = list(result.fetchall())
#                             if r and r[0][0] not in ("", None, " ", "Success"):
#                                 error_message = str(r[0][0])
#                                 # Ensure proper logging for errors
#                                 cursor.callproc('stp_site_error_log', [upload_for, company_id, file_name, datetime.now().date(), error_message, checksum_id, site_id1])
#                                 error_count += 1
                    
#                     if not row_error_found:
#                         # Check if the row is an update or a new insert
#                         cursor.callproc('+', params)
#                         for result in cursor.stored_results():
#                             update_result = result.fetchone()
#                             if update_result is not None:  # Check if there's a result
#                                 if update_result[0] == "Updated":
#                                     update_count += 1  # Increment update count
#                                     row_was_updated = True
#                                 elif update_result[0] == "Success":
#                                     success_count += 1
                    
#             if entity == 'um':
#                 checksum_msg = (
#                     f"Total Columns Processed in each row: {total_columns}, Total Rows Processed: {total_rows}, Successful Entries: {success_count}"
#                     f"{f', Update Entries: {update_count}' if update_count > 0 else ''}"
#                     f"{f', Errors: {error_count}' if error_count > 0 else ''}"
#                 )

#                 cursor.callproc('stp_update_checksum', (
#                     upload_for, company_id, '', str(datetime.now().month), str(datetime.now().year),
#                     file_name, checksum_msg, error_count, update_count, checksum_id, employee_id1
#                 ))

#                 if error_count == 0 and update_count == 0 and success_count > 0:
#                     messages.success(request, "All data uploaded successfully!")
#                 elif error_count == 0 and success_count == 0 and update_count > 0:
#                     messages.warning(request, "All data updated successfully!")
#                 else:
#                     messages.warning(request,
#                         f"The upload processed {total_columns} columns, {total_rows} rows, resulting in {success_count} successful entries"
#                         f"{f', {update_count} updates' if update_count > 0 else ''}, and {error_count} errors; "
#                         "please check the error logs for details."
#                     )

#             elif entity == 'sm':
#                 checksum_msg = (
#                     f"Total Columns Processed in each row: {total_columns}, Total Rows Processed: {total_rows}, Successful Entries: {success_count}"
#                     f"{f', Update Entries: {update_count}' if update_count > 0 else ''}"
#                     f"{f', Errors: {error_count}' if error_count > 0 else ''}"
#                 )

#                 cursor.callproc('stp_update_checksum', (
#                     upload_for, company_id, '', str(datetime.now().month), str(datetime.now().year),
#                     file_name, checksum_msg, error_count, update_count, checksum_id
#                 ))

#                 if error_count == 0 and update_count == 0 and success_count > 0:
#                     messages.success(request, "All data uploaded successfully!")
#                 elif error_count == 0 and success_count == 0 and update_count > 0:
#                     messages.warning(request, "All data updated successfully!")
#                 else:
#                     messages.warning(request,
#                         f"The upload processed {total_columns} columns, {total_rows} rows, resulting in {success_count} successful entries"
#                         f"{f', {update_count} updates' if update_count > 0 else ''}, and {error_count} errors; "
#                         "please check the error logs for details."
#                     )

                   
#         except Exception as e:
#             tb = traceback.extract_tb(e.__traceback__)
#             fun = tb[0].name
#             cursor.callproc("stp_error_log", [fun, str(e), user])  
#             messages.error(request, 'Oops...! Something went wrong!')
#             m.commit()   
#         finally:
#             cursor.close()
#             m.close()
#             Db.closeConnection()
#             if entity=='um':
#                 return redirect(f'/masters?entity=em&type=i')
#             else:
#                 return redirect(f'/masters?entity={entity}&type=i')
            
@login_required  
def upload_excel_cm(request):

    if request.method == 'POST' and request.FILES.get('excelFile'):
        excel_file = request.FILES['excelFile']
        file_name = excel_file.name
        df = pd.read_excel(excel_file)
        total_rows = len(df)
        update_count = error_count = success_count = 0
        checksum_id = None
        r=None
        global user
        user  = request.session.get('user_id', '')
        try:
            Db.closeConnection()
            m = Db.get_connection()
            cursor = m.cursor()
            entity = request.POST.get('entity', '')
            type = request.POST.get('type', '')
            company_id1 = request.POST.get('company_id', None)
            cursor.callproc("stp_get_masters", [entity, type, 'sample_xlsx',user])
            for result in cursor.stored_results():
                columns = [col[0] for col in result.fetchall()]
            if not all(col in df.columns for col in columns):
                messages.error(request, 'Oops...! The uploaded Excel file does not contain the required columns.!')
                return redirect(f'/masters?entity={entity}&type={type}')
            upload_for = {'em': 'employee master','sm': 'site master','cm': 'company master','r': 'roster'}[entity]
            cursor.callproc('stp_insert_checksum', (upload_for,company_id1,str(datetime.now().month),str(datetime.now().year),file_name))
            for result in cursor.stored_results():
                c = list(result.fetchall())
            checksum_id = c[0][0]

            for index,row in df.iterrows():
                params = tuple(str(row.get(column, '')) for column in columns)
                cursor.callproc('stp_insert_company_master', params)
                for result in cursor.stored_results():
                    r = list(result.fetchall())
                if r[0][0] not in ("success", "updated"):
                    cursor.callproc('stp_insert_error_log_cm', [upload_for,company_id1,file_name,datetime.now().date(),str(r[0][0]),checksum_id,])
                if r[0][0] == "success": success_count += 1 
                elif r[0][0] == "updated": update_count += 1  
                else: error_count += 1
            checksum_msg = f"Total Rows Processed: {total_rows}, Successful Entries: {success_count}" f"{f', Updates: {update_count}' if update_count > 0 else ''}" f"{f', Errors: {error_count}' if error_count > 0 else ''}"
            cursor.callproc('stp_update_checksum', (upload_for,company_id1,'',str(datetime.now().month),str(datetime.now().year),file_name,checksum_msg,error_count,update_count,checksum_id))
            if error_count == 0 and update_count == 0 and success_count > 0:
                messages.success(request, f"All data uploaded successfully!.")
            elif error_count == 0 and success_count == 0 and update_count > 0:
                messages.warning(request, f"All data updated successfully!.")
            else:messages.warning(request, f"The upload processed {total_rows} rows, resulting in {success_count} successful entries"  f"{f', {update_count} updates' if update_count > 0 else ''}" f", and {error_count} errors; please check the error logs for details.")
                   
        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name
            cursor.callproc("stp_error_log", [fun, str(e), user])  
            messages.error(request, 'Oops...! Something went wrong!')
            m.commit()   
        finally:
            cursor.close()
            m.close()
            Db.closeConnection()
            return redirect(f'/masters?entity={entity}&type=i')

def get_access_control(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    company = []
    worksite = []
    global user
    user  = request.session.get('user_id', '')
    try:
        if request.method == "POST":
            type1 = request.POST.get('type1','')
            type = request.POST.get('type','')
            ur = request.POST.get('ur', '')
            cursor.callproc("stp_get_access_control_val", [type,ur,'company'])
            for result in cursor.stored_results():
                company = list(result.fetchall())
            cursor.callproc("stp_get_access_control_val", [type,ur,'worksite'])
            for result in cursor.stored_results():
                worksite = list(result.fetchall())
            if type1 == 'worksites':
                company_id = request.POST.getlist('company_id','')
                company_ids = ','.join(company_id)
                cursor.callproc("stp_get_access_control_val", [type1,company_ids,'worksites'])
                for result in cursor.stored_results():
                    worksite = list(result.fetchall())
                response = {
                'result': 'success',
                'worksite':worksite,
                'company': company,
                'worksite': worksite,
                }
            else:     
                response = {
                'result': 'success',
                'worksite':worksite,
                'company': company,
                }

        # Return JSON response
            return JsonResponse(response)
        else: response = {'result': 'fail', 'message': 'Invalid request method'}

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        cursor.callproc("stp_error_log", [tb[0].name, str(e), user])
        print(f"error: {e}")
        response = {'result': 'fail', 'message': 'Something went wrong!'}

    finally:
        cursor.close()
        m.close()
        Db.closeConnection()
        return JsonResponse(response)

    
class SlotDataAPIView(APIView):
    # Ensure the user is authenticated using JWT
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        # Extract user ID from the JWT token
        user = request.user  # This will get the user from the JWT token

        # Call the function to get the roster data
        slot_data = self.get_slot_data(user.id)
        Log.objects.create(log_text=f"Fetched user by ID: {user.id}")

        return Response(slot_data)

    def get_slot_data(self, user_id):
        # Close any previous connections and get a new one
        Db.closeConnection()
        m = Db.get_connection()
        cursor = m.cursor()

        # Step 1: Retrieve the user
        try:
            user = CustomUser.objects.get(id=user_id)
            phone_number = user.phone
        except CustomUser.DoesNotExist:
            return {'error': 'User not found'}

        try:
            employee = sc_employee_master.objects.get(mobile_no=phone_number)
            employee_id = employee.employee_id
            company_idd = employee.company_id.company_id
        except sc_employee_master.DoesNotExist:
            return {'error': 'Employee not found'}

        # Step 3: Retrieve all UserSlotDetails entries for the given employee and company
        try:
            user_slot_details = UserSlotDetails.objects.filter(employee_id=employee_id)
            user_slot_data = UserSlotDetailsSerializer(user_slot_details, many=True).data

            employee_data = sc_employee_master.objects.filter(employee_id=employee_id).first()
            if employee_data:
                mobile_no = employee_data.mobile_no
            else:
                mobile_no = None 
 
            user_alloted_count = len(user_slot_data)

            name = employee_data.employee_name

            user_attendance_details = slot_attendance_details.objects.filter(employee_id=employee_id)
            user_attendance_data = UserSlotAttendedSerializer(user_attendance_details, many=True).data

            user_attendance_count = len(user_attendance_data)

            designation_ids = employee_designation.objects.filter(employee_id=employee_id).values_list('designation_id', flat=True)
            site_ids = employee_site.objects.filter(employee_id=employee_id).values_list('site_id', flat=True)

            slot_details = SlotDetails.objects.filter(
                designation_id__in=designation_ids,
                site_id__in=site_ids
            )

            current_date = timezone.now().date()

            filtered_slot_details = slot_details.filter(shift_date__gt=current_date).order_by('shift_date')

            slot_details_list = SlotListDetailsSerializer(filtered_slot_details, many=True).data

            slot_ids = filtered_slot_details.values_list('slot_id', flat=True)

            slot_id_counts = UserSlotDetails.objects.filter(
                slot_id__in=slot_ids,
                employee_id=employee_id
            ).values('slot_id').annotate(
                count=Count('id')
            ).values('slot_id', 'id', 'employee_id', 'count')

            slot_count_dict = {slot_id: {'count': 0, 'id': 0, 'employee_id': employee_id} for slot_id in slot_ids}

            # Update counts based on slot_id_counts
            for item in slot_id_counts:
                slot_id = item['slot_id']
                if slot_id in slot_count_dict:
                    slot_count_dict[slot_id] = {
                        'count': item['count'],
                        'id': item['id'],
                        'employee_id': item['employee_id']
                    }

            slot_count_list = [
                {
                    'employeeId': data['employee_id'],
                    'slotId': slot_id,
                    'id': data['id'],
                    'count': data['count']
                }
                for slot_id, data in slot_count_dict.items()
            ]

            print(slot_count_list)

            wage_data_queryset = salary_generated_log.objects.filter(employee_id=employee_id)
            wage_data_count = wage_data_queryset.count()

            utr_data_queryset = PayoutDetails.objects.filter(employee_id=employee_id)

            # Get the count of records
            utr_count = utr_data_queryset.count()

            # Get the data as a list of dictionaries
            utr_data = list(utr_data_queryset.values())


            wage_data = salary_generated_log.objects.filter(employee_id=employee_id)
            salary_data = []

            if wage_data.exists():
                for wage in wage_data:
                    slot_id = wage.slot_id.slot_id
                    filtered_salaries = daily_salary.objects.filter(
                        employee_id=employee_id,
                        slot_id=slot_id,
                        element_name__in=["Net Salary"]
                    )
                    if filtered_salaries.exists():
                        salary_data.extend(DailySalarySerializer(filtered_salaries, many=True).data)

            return {
                'name':name,
                'slot_alloted_count': user_alloted_count,
                'slot_alloted_list': list(user_slot_data),
                'user_attendance_count':user_attendance_count,
                'user_attendance_list':list(user_attendance_data),
                'slot_details_list':list(slot_details_list),
                'slot_count_list':slot_count_list,
                'employee_id':employee_id,
                'mobile_no':mobile_no,
                'wage_count':wage_data_count,
                'utr_count':utr_count,
                'salary_data':salary_data,
                'utr_data':utr_data
            }
        
        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            cursor.callproc("stp_error_log", [tb[0].name, str(e), user_id])
            print(f"error: {e}")
            return {'result': 'fail', 'message': 'Something went wrong!'}


        
class confirm_schedule(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]
    # authentication_classes = []

    def post(self, request):
        try:
            data = request.data
            roster_id = request.data.get('id')
            confirmation = request.data.get('confirmation') == '1'
            user = request.user
            roster = sc_roster.objects.get(id=roster_id)
            roster.confirmation = confirmation
            roster.updated_at = timezone.now()
            roster.confirmation_date = timezone.now()
            roster.updated_by = user
            roster.save()
            ser = ScRosterSerializer(roster)

            return Response({'success': 'Confirmation updated successfully.','data':ser.data,'con':confirmation}, status=200)
        except sc_roster.DoesNotExist:
            return Response({'error': 'Roster not found.'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)



class confirm_notification(APIView):
    # Ensure the user is authenticated using JWT
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]
    # authentication_classes = []

    def post(self, request):
        try:
            data= request.data
            notification_id = request.data.get('id')
            confirmation = request.data.get('confirmation') == '1'
            user = request.user
            notification = notification_log.objects.get(id=notification_id)
            notification.notification_opened = timezone.now()
            notification.updated_at = timezone.now()
            notification.updated_by = user
            notification.save()

            return Response({'success': 'Confirmation updated successfully.'}, status=200)
        except notification_log.DoesNotExist:
            return Response({'error': 'notification_log not found.'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)
        
@login_required
def slot_details(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    user_id = request.session.get('user_id', '')

    try:
        if request.method == 'GET':
            slot_id = request.GET.get('slot_id', '')
            type = request.GET.get('type', '')

            cursor.callproc("stp_get_assigned_company", [user_id])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())

            # cursor.callproc("stp_get_userwise_dropdown", [user_id,'site'])
            # for result in cursor.stored_results():
            #     site_names = list(result.fetchall())
            
            cursor.callproc("stp_get_dropdown_values", ['designation'])
            for result in cursor.stored_results():
                designation = list(result.fetchall())

            if slot_id == "0":
                context = {'slot_id': slot_id,'company_names': company_names,'designation':designation,'type': type}


        elif request.method == 'POST':
            slot_id = request.POST.get('slot_id', '')
            type = 'shift'
            company_id = request.POST.get('company_id', '')
            site_id = request.POST.get('worksite', '')  

            cursor.callproc("stp_get_dropdown_values", ['designation'])
            for result in cursor.stored_results():
                designation = list(result.fetchall())

            context= {'slot_id':slot_id,'company_id':company_id, 'site_id':site_id,'type':type,'designation':designation}
        
            # messages.success(request, "Slot successfully created!")
            

    except Exception as e:
        tb = traceback.format_exc()  # Capture the full traceback
        cursor.callproc("stp_error_log", [tb, str(e), str(user_id)])  # Log the error with the traceback and user ID
        print(f"error: {e}")
        return JsonResponse({'result': 'fail', 'message': 'Something went wrong!'})

    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        return render(request, 'Master/slot/slot_details.html', context)


@login_required
def post_slot_details(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    user_idd = request.session.get('user_id', '')
    user = CustomUser.objects.get(id=user_idd)  

    try:
        # Load additional shift data from JSON input
        shifts_data = json.loads(request.POST.get('shifts', '[]'))  
        shifts_data2 = json.loads(request.POST.get('shifts2', '[]')) 

        # Gather main shift data from POST request
        slot_name = request.POST.get('slot_name', '')
        description = request.POST.get('description', '')

        # Convert shift_date to a date object
        shift_date_str = request.POST.get('shift_date', '')
        shift_date = datetime.strptime(shift_date_str, '%Y-%m-%d').date() if shift_date_str else None

        start_time = request.POST.get('start_time', '')
        end_time = request.POST.get('end_time', '')
        designation = request.POST.get('designation')
        night_shift = request.POST.get('night_shift', '0')
        company_id = request.POST.get('company_id', '')
        site_id1 = request.POST.get('site_id', '')

        if not shift_date and not shifts_data and not shifts_data2:
            return JsonResponse({'status': 'error', 'message': 'No shift data received.'})

        def insert_setting_master(slot_detail):
            # Create SettingMaster entry
            setting_master = SettingMaster.objects.create(
                noti_start_time=timezone.now().date(),
                noti_end_time=slot_detail.shift_date - timedelta(days=1),
                interval=8,
                no_of_notification=2,
                no_of_employee=9999,
                slot_id=slot_detail,
                created_by=user
            )
            # Update SlotDetails with the SettingMaster ID
            slot_detail.setting_id = get_object_or_404(SettingMaster, id=setting_master.id)
            slot_detail.save()

        # Save the main shift if all necessary data is provided
        if shift_date and start_time and end_time:
            initial_shift_detail = SlotDetails(
                slot_name=slot_name,
                slot_description=description,
                shift_date=shift_date,
                start_time=start_time,
                end_time=end_time,
                designation_id=get_object_or_404(designation_master, designation_id=designation),
                night_shift=bool(int(night_shift)), 
                company_id=company_id,
                site_id=get_object_or_404(sit, site_id=site_id1),
                created_by=user
            )
            initial_shift_detail.save()
            insert_setting_master(initial_shift_detail)

        # Process extra shifts (shifts_data)
        for shift in shifts_data:
            if shift.get('start_time') and shift.get('end_time'):
                shift_detail_add = SlotDetails(
                    slot_name=slot_name,
                    slot_description=description,
                    shift_date=shift_date,
                    designation_id=get_object_or_404(designation_master, designation_id=designation),
                    start_time=shift.get('start_time'),
                    end_time=shift.get('end_time'),
                    night_shift=bool(int(shift.get('night_shift', '0'))), 
                    company_id=company_id,
                    site_id=get_object_or_404(sit, site_id=site_id1),
                    created_by=user
                )
                shift_detail_add.save()
                insert_setting_master(shift_detail_add)

        # Process multiple shifts (shifts_data2)
        for shift2 in shifts_data2:
            if shift2.get('shiftDate') and shift2.get('startTime') and shift2.get('endTime'):
                shift2_date_str = shift2.get('shiftDate')
                shift2_date = datetime.strptime(shift2_date_str, '%Y-%m-%d').date()
                shift_detail_add = SlotDetails(
                    slot_name=shift2.get('new_slot_name'),
                    slot_description=shift2.get('new_description'),
                    designation_id=get_object_or_404(designation_master, designation_id=shift2.get('new_designation')),
                    shift_date=shift2_date,
                    start_time=shift2.get('startTime'),
                    end_time=shift2.get('endTime'),
                    night_shift=bool(int(shift2.get('newNightShift', '0'))), 
                    company_id=company_id,
                    site_id=get_object_or_404(sit, site_id=site_id1),
                    created_by=user
                )
                shift_detail_add.save()
                insert_setting_master(shift_detail_add)

            for additional_shift in shift2.get('shiftTimes', []):
                if additional_shift.get('newStartTime') and additional_shift.get('newEndTime'):
                    shift_detail_additional = SlotDetails(
                        slot_name=shift2.get('new_slot_name'),
                        slot_description=shift2.get('new_description'),
                        designation_id=get_object_or_404(designation_master, designation_id=shift2.get('new_designation')),
                        shift_date=shift2_date,
                        start_time=additional_shift.get('newStartTime'),
                        end_time=additional_shift.get('newEndTime'),
                        night_shift=bool(int(additional_shift.get('newNightShift', '0'))),  
                        company_id=company_id,
                        site_id=get_object_or_404(sit, site_id=site_id1),
                        created_by=user
                    )
                    shift_detail_additional.save()
                    insert_setting_master(shift_detail_additional)

        return JsonResponse({'success': True, 'redirect_url': '/masters?entity=sd&type=i'})

    except Exception as e:
        tb = traceback.format_exc()
        cursor.callproc("stp_error_log", [tb, str(e), str(user_idd)])
        print(f"error: {e}")
        
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()



@login_required 
def setting_master(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    user_id  = request.session.get('user_id', '')
    user_idd = CustomUser.objects.get(id=user_id)
    try:
        if request.method == 'GET':
            slot_id = request.GET.get('slot_id', '')
            slot_idd = decrypt_parameter(slot_id)
            type = request.GET.get('type', '')
            # setting_id = request.GET.get('setting_id', '')

            try:
                settings_data = get_object_or_404(SettingMaster, slot_id=slot_idd)
                context = {'settings_data': settings_data, 'type': type}
            except Http404:
                context = {'slot_id':slot_id,'type': type} 
            
        elif request.method == 'POST':
            setting_id = request.POST.get('setting_id1', '')
            # slot_id = request.POST.get('slot_id', '')
            # slot_idd = decrypt_parameter(slot_id)
            # slot_id1 = get_object_or_404(SlotDetails, slot_id=slot_idd),
            if setting_id:
                notification = get_object_or_404(SettingMaster, id=setting_id)
                notification.noti_start_time = request.POST.get('notification_start_time', '')
                notification.noti_end_time = request.POST.get('notification_end_time', '')
                notification.no_of_notification = request.POST.get('number_of_notifications', '')
                notification.interval = request.POST.get('notification_interval_hours', '')
                notification.no_of_employee = request.POST.get('employee_count', '')
                notification.updated_by = user_idd
                notification.save()

                messages.success(request, "Slot Settings successfully updated!")
            else:
                notification = SettingMaster(
                noti_start_time= request.POST.get('notification_start_time', ''),
                noti_end_time=request.POST.get('notification_end_time', ''),
                no_of_notification=request.POST.get('number_of_notifications', ''),
                interval = request.POST.get('notification_interval_hours',''),
                no_of_employee = request.POST.get('employee_count',''),
                slot_id = get_object_or_404(slot, slot_id=decrypt_parameter(request.POST.get('slot_id', ''))),
                created_by=user_idd
                )
                notification.save()
                new_setting_id = notification.id

                # Update SlotDetails with the new SettingMaster ID where slot_id matches
                slot_id = decrypt_parameter(request.POST.get('slot_id', ''))
                SlotDetails.objects.filter(slot_id=slot_id).update(setting_id=new_setting_id)

                messages.success(request, "Slot Settings successfully saved and SlotDetails updated!")




    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        cursor.callproc("stp_error_log", [tb[0].name, str(e), user])
        print(f"error: {e}")
        response = {'result': 'fail', 'message': 'Something went wrong!'}
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method == 'GET':
            return render(request, 'Master/slot/slot_details.html', context)
        elif request.method == 'POST':
            new_url = f'/masters?entity=sd&type=i'
            return redirect(new_url) 

@login_required  
def edit_slot_details(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    user_id  = request.session.get('user_id', '')
    user_idd = CustomUser.objects.get(id=user_id)
        

    try:
        if request.method == 'GET':
            cursor.callproc("stp_get_assigned_company", [user_id])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())

            # cursor.callproc("stp_get_graph_dropdown", [user_id,'site'])
            # for result in cursor.stored_results():
            #     site_names = list(result.fetchall())

        

            cursor.callproc("stp_get_dropdown_values", ['designation'])
            for result in cursor.stored_results():
                designation = list(result.fetchall())
            slot_id = request.GET.get('slot_id')
            slot_idd = decrypt_parameter(slot_id)
            slot_data = get_object_or_404(SlotDetails, slot_id=slot_idd)
            print(slot_data.company.company_id)

            cursor.callproc("stp_get_slot_master_worksite", [user_id,slot_data.company.company_id])
            for result in cursor.stored_results():
                site_names = list(result.fetchall())
          
            context = {
                'slot_data': slot_data,
                'company_names':company_names,
                'designation':designation,
                'site_name':site_names
            }
        elif request.method == 'POST':
            slot_id = request.POST.get('slot_id')
            slot_data = get_object_or_404(SlotDetails, slot_id=slot_id)
            company_id = request.POST.get('company_id')
            slot_name = request.POST.get('slot_name')
            description = request.POST.get('description')
            shift_date  = request.POST.get('shift_date')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            designation_id=request.POST.get('designation')
            night_shift = 1 if request.POST.get('night_shift') == 'on' else 0

            current_date = date.today()  # Get the current date

            setting = get_object_or_404(SettingMaster, slot_id=slot_id)
            noti_start_time = setting.noti_start_time  

            if current_date > noti_start_time:  # Compare current date with noti_start_time
                messages.error(request, "You Cannot Update Data as Notification Time has Started.")
                return
                
            slot_data.company_id = company_id
            slot_data.site_id = get_object_or_404(sit, site_id=request.POST.get('site_id',''))
            slot_data.designation_id = get_object_or_404(designation_master, designation_id=designation_id)
            slot_data.slot_name = slot_name
            slot_data.slot_description = description
            slot_data.shift_date = shift_date
            slot_data.start_time = start_time
            slot_data.end_time = end_time
            slot_data.night_shift = night_shift
            slot_data.updated_by = user_idd

            # Save the updated slot details to the database
            slot_data.save()

            # Show success message and redirect
            messages.success(request, "Slot details updated successfully.")
            
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        cursor.callproc("stp_error_log", [tb[0].name, str(e), user])
        print(f"error: {e}")
        response = {'result': 'fail', 'message': 'Something went wrong!'}
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method=="GET":
            return render(request, 'Master/slot/edit_slot_details.html', context)
        if request.method=="POST":
            new_url = f'/masters?entity=sd&type=i'
            return redirect(new_url) 
        

# def delete_slot(request):
#     Db.closeConnection()  
#     m = Db.get_connection()
#     cursor = m.cursor()

#     try:
#         slot_id = request.POST.get('slot_id')
#         slot_idd = decrypt_parameter(slot_id)

#         slots_to_delete = SlotDetails.objects.filter(slot_id=slot_idd)
#         setting = get_object_or_404(SettingMaster, slot_id=slot_idd)
#         noti_start_time = setting.noti_start_time 

#         current_date = date.today()

#         if current_date >= noti_start_time:  
#             return JsonResponse({'success': False, 'message': 'Slot Cannot be deleted,Because Notification time has begun!'})

#         if slots_to_delete.exists():
#             slots_to_delete.delete()
#             setting.delete()
#             return JsonResponse({'success': True, 'message': 'Slot  and its setttings successfully deleted!'})
#         else:
#             return JsonResponse({'success': False, 'message': 'No slots found with the specified slot_id.'})

#     except Exception as e:
#         tb = traceback.extract_tb(e.__traceback__)
#         cursor.callproc("stp_error_log", [tb[0].name, str(e), request.user.username])  

#         return JsonResponse({'success': False, 'message': 'An error occurred while deleting the slot.'})

def delete_slot(request):
    Db.closeConnection()  
    m = Db.get_connection()
    cursor = m.cursor()

    try:
        # Decrypt slot_id from POST request
        slot_id = request.POST.get('slot_id')
        slot_idd = decrypt_parameter(slot_id)

        # Fetch the slot to delete and check its settings
        slot_to_delete = get_object_or_404(SlotDetails, slot_id=slot_idd)
        setting = get_object_or_404(SettingMaster, slot_id=slot_idd)
        
        noti_start_time = setting.noti_start_time  # Get notification start time from settings
        current_date = date.today()

        # If notification has already started, do not delete the slot
        if current_date >= noti_start_time:
            return JsonResponse({'success': False, 'message': 'Slot cannot be deleted, because notification time has begun!'})

        # Delete the specific slot and setting if it exists and notification time has not started
        slot_to_delete.delete()
        setting.delete()

        return JsonResponse({'success': True, 'message': 'Slot and its settings successfully deleted!'})

    except Exception as e:
        # Log the error if any occurs
        tb = traceback.extract_tb(e.__traceback__)
        cursor.callproc("stp_error_log", [tb[0].name, str(e), request.user.username])  

        return JsonResponse({'success': False, 'message': 'An error occurred while deleting the slot.'})



def view_employee(request):
    Db.closeConnection()  
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        if request.method == "GET":
            id = request.GET.get('id', '')

            # Initialize context variable before starting to populate it
            context = {
                'id': id,
                'employee_status': [],
                'employee_status_id': '',
                'site_name': [],
                'gender': [],
                'company_names': [],
                'state_names': [],
                'designation_name': [],
                'employee_id': None,
                'employee_name': None,
                'mobile_no': None,
                'email': None,
                'gender_value': None,
                'state_id_value': None,
                'city': None,
                'address': None,
                'pincode': None,
                'account_holder_name': None,
                'account_no': None,
                'bank_name': None,
                'branch_name': None,
                'ifsc_code': None,
                'pf_no': None,
                'uan_no': None,
                'esic': None,
                'employment_status_id': None,
                'handicapped_value': None,
                'is_active': None,
                'company_id_value': None,
                'selected_designation': [],
                'selected_site': []
            }

            # Populate dropdown values
            cursor.callproc("stp_get_employee_status")
            for result in cursor.stored_results():
                context['employee_status'] = list(result.fetchall())

            cursor.callproc("stp_get_dropdown_values", ('site',))
            for result in cursor.stored_results():
                context['site_name'] = list(result.fetchall())

            cursor.callproc("stp_get_dropdown_values", ('company',))
            for result in cursor.stored_results():
                context['company_names'] = list(result.fetchall())

            cursor.callproc("stp_get_dropdown_values", ('states',))
            for result in cursor.stored_results():
                context['state_names'] = list(result.fetchall())

            cursor.callproc("stp_get_dropdown_values", ('gender',))
            for result in cursor.stored_results():
                context['gender'] = list(result.fetchall())

            cursor.callproc("stp_get_dropdown_values", ('designation',))
            for result in cursor.stored_results():
                context['designation_name'] = list(result.fetchall())

           

            # Otherwise, we are editing an existing employee
            id1 = request.GET.get('id', '')
            id = decrypt_parameter(id1)

            cursor.callproc("stp_edit_employee_master", (id,))
            for result in cursor.stored_results():
                data = result.fetchall()[0]
                context.update({
                    'employee_id': data[0],
                    'employee_name': data[1],
                    'mobile_no': data[2],
                    'email': data[3],
                    'gender_value': data[4],
                    'state_id_value': data[5],
                    'city': data[6],
                    'address': data[7],
                    'pincode': data[8],
                    'account_holder_name': data[9],
                    'account_no': data[10],
                    'bank_name': data[11],
                    'branch_name': data[12],
                    'ifsc_code': data[13],
                    'pf_no': data[14],
                    'uan_no': data[15],
                    'esic': data[16],
                    'employment_status_id': data[17],
                    'handicapped_value': data[18],
                    'is_active': data[19],
                    'company_id_value': data[20],
                })

                # Get the employee's designation and site info
                employee_id = context['employee_id']
                designation_list = employee_designation.objects.filter(employee_id=employee_id)
                context['selected_designation'] = json.dumps(list(map(str, designation_list.values_list('designation_id', flat=True))))

                site_list = employee_site.objects.filter(employee_id=employee_id)
                context['selected_site'] = json.dumps(list(map(str, site_list.values_list('site_id', flat=True))))

    except Exception as e:
        print(f"Error: {e}")  # Temporary for debugging
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])  
        messages.error(request, 'Oops...! Something went wrong!')

    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()

        if request.method == "GET":
            return render(request, "Master/master/employee_view.html", context)
        elif request.method == "POST":  
            return redirect(f'/masters?entity=em&type=i')
        

@login_required        
def designation_master1(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    global user
    
    # user_id = request.session.get('user_id', '')
    # user = CustomUser.objects.get(id=user_id)
    try:
        
        if request.method == "GET":
           
            designation_id = request.GET.get('designation_id', '')
            if designation_id == "0":
                if request.method == "GET":
                    context = {'designation_id':designation_id}

            else:
                designation_id = request.GET.get('designation_id', '')
                designation_id = decrypt_parameter(designation_id)
                cursor.callproc("stp_designationedit", (designation_id,))
                for result in cursor.stored_results():
                    data = result.fetchall()[0]  
                    context = {
                        'designation_id': data[0],
                        'designation_name': data[1],
                        'is_active':data[2]
                    }

        if request.method == "POST" :
            id = request.POST.get('designation_id', '')
            if id == '0':
                designation_name = request.POST.get('designation_name', '')
                # is_active = request.POST.get('is_active', '') 
                params = [
                    designation_name,
                    # is_active  
                ]
                cursor.callproc("stp_designationinsert", params)
                for result in cursor.stored_results():
                        datalist = list(result.fetchall())
                if datalist[0][0] == "success":
                    messages.success(request, 'Data successfully entered !')
                else: messages.error(request, datalist[0][0])
            else:
                designation_id = request.POST.get('designation_id', '')
                designation_name = request.POST.get('designation_name', '')
                # is_active1 = request.POST.get('is_active', '')
                is_active = 1 if request.POST.get('is_active') == 'on' else 0 
                            
                params = [designation_id,designation_name,is_active]    
                cursor.callproc("stp_designationupdate",params) 
                messages.success(request, "Data successfully Updated!")

    except Exception as e:
        print(f"Error: {e}")  # Temporary for debugging
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])  
        messages.error(request, 'Oops...! Something went wrong!')

    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()

        if request.method == "GET":
            return render(request, "Master/master/designation_master.html", context)
        elif request.method == "POST":  
            return redirect(f'/masters?entity=dm&type=i')

def view_designation(request):
    Db.closeConnection()  
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        if request.method == "GET":
            designation_id = request.GET.get('designation_id', '')

            # Initialize context variable before starting to populate it
            context = {
                'designation_id': None,
                'designation_name': None,
                'is_active': None
            }

            # Otherwise, we are editing an existing employee
            designation_id1 = request.GET.get('designation_id', '')
            designation_id = decrypt_parameter(designation_id1)

            cursor.callproc("stp_designationedit", (designation_id,))
            for result in cursor.stored_results():
                data = result.fetchall()[0]
                context.update({
                    'designation_id': data[0],
                    'designation_name': data[1],
                    'is_active': data[2],
                })


    except Exception as e:
        print(f"Error: {e}")  # Temporary for debugging
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])  
        messages.error(request, 'Oops...! Something went wrong!')

    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()

        if request.method == "GET":
            return render(request, "Master/master/designation_view.html", context)
        elif request.method == "POST":  
            return redirect(f'/masters?entity=dm&type=i')

@login_required
def employee_upload(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    pre_url = request.META.get('HTTP_REFERER')
    header, data = [], []
    entity, type, name = '', '', ''
    global user
    user  = request.session.get('user_id', '')
    try:
         
        if request.method=="GET":
            entity = request.GET.get('entity', '')
            type = request.GET.get('type', '')
            
            cursor.callproc("stp_get_company_site_name",[user])
            for result in cursor.stored_results():
                site_name = list(result.fetchall())
            cursor.callproc("stp_get_dropdown_values",['designation'])
            for result in cursor.stored_results():
                designation_name = list(result.fetchall())
            cursor.callproc("stp_get_assigned_company",[user])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())
            cursor.callproc("stp_get_dropdown_values",['states'])
            for result in cursor.stored_results():
                state_name = []
                state_name = list(result.fetchall())
                

        if request.method=="POST":
            entity = request.POST.get('entity', '')
            type = request.POST.get('type', '')
            if entity == 'r' and type == 'ed':
                ids = request.POST.getlist('ids[]', '')
                shifts = request.POST.getlist('shifts[]', '')
                for id,shift in zip(ids, shifts):
                    cursor.callproc("stp_post_roster",[id,shift])
                    for result in cursor.stored_results():
                        datalist = list(result.fetchall())
                if datalist[0][0] == "success":
                    messages.success(request, 'Data updated successfully !')
            if entity == 'urm' and (type == 'acu' or type == 'acr'):
                
                created_by = request.session.get('user_id', '')
                ur = request.POST.get('ur', '')
                selected_company_ids = list(map(int, request.POST.getlist('company_id')))
                selected_worksites  = request.POST.getlist('worksite')
                company_worksite_map = {}
                
                if not selected_company_ids or not selected_worksites:
                    messages.error(request, 'Company or worksite data is missing!')
                    return redirect(f'/masters?entity={entity}&type=urm')
                if type not in ['acu', 'acr'] or not ur:
                    messages.error(request, 'Invalid data received.')
                    return redirect(f'/masters?entity={entity}&type=urm')
                
                cursor.callproc("stp_get_company_worksite",[",".join(request.POST.getlist('company_id'))])
                for result in cursor.stored_results():
                    company_worksites  = list(result.fetchall())
                    
                for company_id, worksite_name in company_worksites:
                    if company_id not in company_worksite_map:
                        company_worksite_map[company_id] = []
                    company_worksite_map[company_id].append(worksite_name)
                
                filtered_combinations = []
                for company_id in selected_company_ids:
                    valid_worksites = company_worksite_map.get(company_id, [])
                    # Filter worksites that were actually selected by the user
                    selected_valid_worksites = [ws for ws in selected_worksites if ws in valid_worksites]
                    filtered_combinations.extend([(company_id, ws) for ws in selected_valid_worksites])
                    
                cursor.callproc("stp_delete_access_control",[type,ur])
                r=''
                for company_id, worksite in filtered_combinations:
                    cursor.callproc("stp_post_access_control",[type,ur,company_id,worksite,created_by])
                    for result in cursor.stored_results():
                            r = list(result.fetchall())
                type='urm'
                if r[0][0] == "success":
                    messages.success(request, 'Data updated successfully !')
                
            else : messages.error(request, 'Oops...! Something went wrong!')
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method=="GET":
            return render(request,'Master/employee_upload.html', {'entity':entity,'type':type,'name':name,'company_names':company_names,'state_name':state_name,'site_name':site_name,'designation_name':designation_name,'pre_url':pre_url})
        elif request.method=="POST":  
            new_url = f'/masters1?entity={entity}&type={type}'
            return redirect(new_url)      

@login_required
def worksite_upload(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    pre_url = request.META.get('HTTP_REFERER')
    header, data = [], []
    entity, type, name = '', '', ''
    global user
    user  = request.session.get('user_id', '')
    try:
         
        if request.method=="GET":
            entity = request.GET.get('entity', '')
            type = request.GET.get('type', '')
            
            cursor.callproc("stp_get_userwise_dropdown",[user,'company'])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())
                  # Fetching combined city and state data using stp_city_state
            cursor.callproc("stp_city_state")
            for result in cursor.stored_results():
                city_state_data = list(result.fetchall())
                
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method=="GET":
            return render(request,'Master/worksite_upload.html', {'entity':entity,'type':type,'name':name,'header':header,'company_names':company_names,'city_state_data':city_state_data,'data':data,'pre_url':pre_url})
        elif request.method=="POST":  
            new_url = f'/masters1?entity={entity}&type={type}'
            return redirect(new_url)  


class EmployeeData(APIView):
    # Ensure the user is authenticated using JWT
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request, *args, **kwargs):
        # Get parameters from the request
        employee_id = request.data['employee_id']
        company_id = request.data['company_id']

        try:
            # Fetch the employee using the provided parameters
            employee = sc_employee_master.objects.filter(employee_id=employee_id, company_id=company_id).first()

            if employee:
                serializer = EmployeeSerializer(employee)
                return Response(serializer.data)
            else:
                return Response({"error": "Employee not found"}, status=404)

        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def put(self, request, *args, **kwargs):
        employee_id = request.data.get('employee_id')
        company_id = request.data.get('company_id')

        try:
            # Fetch the employee using the provided parameters
            employee = sc_employee_master.objects.filter(employee_id=employee_id, company_id=company_id).first()

            if not employee:
                return Response({"error": "Employee not found"}, status=404)
            vv = request.data.get('state_id', employee.state_id)
            cc = get_object_or_404(StateMaster,  state_id=vv)    
            # Update only the fields that are provided in the request
            employee.email = request.data.get('email', employee.email)
            employee.address = request.data.get('address', employee.address)
            employee.city = request.data.get('city', employee.city)
            employee.pincode = request.data.get('pincode', employee.pincode)
            employee.state_id=cc

            employee.save()

            return Response({"success": "Employee data updated successfully"}, status=200)

        except Exception as e:
            return Response({"error": str(e)}, status=500)
class StateName(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        try:
            states = StateMaster.objects.all()  
            serializer = StateMasterSerializer(states, many=True)  
            return Response(serializer.data)  
        except Exception as e:
            return Response({"error": str(e)},  status=404)



def deactivate_slot(request):
    Db.closeConnection()  
    m = Db.get_connection()
    cursor = m.cursor()
    user_id  = request.session.get('user_id', '')
    user_idd = CustomUser.objects.get(id=user_id)
    try:
        if request.method == 'POST':
            slot_id = request.POST.get('slot_id')
            slot_idd = decrypt_parameter(slot_id)
            reason = request.POST.get('reason', '') 

            slot_detail = SlotDetails.objects.get(slot_id=slot_idd)  
            slot_detail.is_active = 0
            slot_detail.message = reason 
            slot_detail.updated_by = user_idd
            slot_detail.save()

            return JsonResponse({'message': 'Slot deactivated successfully.'})
    except SlotDetails.DoesNotExist:
        return JsonResponse({'message': 'Slot ID not found.'}, status=404)
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        cursor.callproc("stp_error_log", [tb[0].name, str(e), request.user.username])
        return JsonResponse({'message': str(e)}, status=500)
    
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()


class post_user_slot(APIView):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]
    
    def post(self, request):
        Db.closeConnection()  
        m = Db.get_connection()
        cursor = m.cursor()
        # Extracting the user id from the session
        try:
            mobileNum = request.data.get('mobileNum')
            user = CustomUser.objects.filter(phone=mobileNum).first()

            if user:
                user_id = user.id  # Access the 'id' of the user
            else:
                user_id = None  # If no user is found, set user_id to None

            employee_id = request.data.get('employee_id')
            slot = request.data.get('slot_id')
            site = request.data.get('site_id')
            slot_id = get_object_or_404(SlotDetails, slot_id=slot)
            company = get_object_or_404(com, company_id=request.data.get('company_id'))
            site = get_object_or_404(sit, site_id=site)
            user = get_object_or_404(CustomUser, id = user_id)

            emp_id1 = get_object_or_404(sc_employee_master, employee_id=employee_id)

                # Creating a new instance of UserSlotDetails and saving it to the database
            user_slot = UserSlotDetails(
                employee_id=employee_id,
                slot_id=slot_id,
                company_id=company,
                site_id=site,
                emp_id = emp_id1,
                created_by = user
            )
            user_slot.save()
            return Response({"message": "User slot details created successfully."}, status=200)
        
        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            cursor.callproc("stp_error_log", [tb[0].name, str(e), request.user.username])
            return JsonResponse({'message': str(e)}, status=500)
        
class delete_user_slot(APIView):
    
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]
    
    def delete(self, request):
        Db.closeConnection()
        m = Db.get_connection()
        cursor = m.cursor()

        try:
            # Retrieve data from the request
            employee_id = request.data.get('employee_id')
            slot_id = request.data.get('slot_id')
            record_id = request.data.get('id')

            # Step 1: Retrieve the UserSlotDetails instance by id
            user_slot_instance = get_object_or_404(UserSlotDetails, id=record_id)

            # Step 2: Prepare the data for insertion into HistUserSlotDetails
            historical_data = {
                'employee_id': user_slot_instance.employee_id,
                'slot_id': user_slot_instance.slot_id,
                'site_id': user_slot_instance.site_id,
                'company_id': user_slot_instance.company_id,
                'created_by': user_slot_instance.created_by,
                'emp_id': user_slot_instance.emp_id,
            }

            # Step 3: Check if a record exists in HistUserSlotDetails for the same slot_id and employee_id
            hist_instance = HistUserSlotDetails.objects.filter(
                slot_id=slot_id,
                employee_id=employee_id
            ).first()

            if hist_instance:
                # If the record exists, update it
                for field, value in historical_data.items():
                    setattr(hist_instance, field, value)
                hist_instance.save()
                message = 'Historical record updated successfully.'
            else:
                # If the record doesn't exist, create a new one
                hist_instance = HistUserSlotDetails(**historical_data)
                hist_instance.save()
                message = 'Historical record created successfully.'

            # Step 4: Delete the record from UserSlotDetails
            user_slot_instance.delete()

            return JsonResponse({'message': message}, status=200)

        except Exception as e:
            # Log the error using a stored procedure
            tb = traceback.extract_tb(e.__traceback__)
            cursor.callproc("stp_error_log", [tb[0].name, str(e), request.user.username])
            return JsonResponse({'message': str(e)}, status=500)


@login_required   
def get_worksites(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    
    try:
        user_id = request.session.get('user_id', '')
        selectedCompany = request.POST.get('selectedCompany','')
        cursor.callproc("stp_get_slot_siteName", [user_id,selectedCompany])
        for result in cursor.stored_results():
            companywise_site_names = list(result.fetchall())

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), request.user.id])
        print(f"error: {e}")
        return JsonResponse({'result': 'fail', 'message': 'something went wrong!'}, status=500)
    
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
    
    return JsonResponse({'companywise_site_names': companywise_site_names}, status=200)

class update_bank_details(APIView):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]
    
    def post(self, request):
        Db.closeConnection()  # Close any previous DB connections
        m = Db.get_connection()  # Get a new connection
        cursor = m.cursor()
        
        # Extract data from the request
        employee_id = request.data.get("employee_id")
        company_id = request.data.get("company_id")
        bank_name = request.data.get("bank_name")
        account_no = request.data.get("account_no")
        holder_name = request.data.get("holder_name")
        ifsc = request.data.get("ifsc")
        branch_name = request.data.get("branch_name")
        phone = request.data.get("phone")

        # emp_id = sc_employee_master.objects.get(employee_id=employee_id).id
        user_id = get_object_or_404(CustomUser, id=CustomUser.objects.get(phone=phone).id)

        bank = BankDetailAPI.objects.first()
        
        # Save data into EmployeeUpdate table using Django ORM
        try:
            employee_update = BankDetails.objects.create(
                employee_id=employee_id,
                company_id=company_id,
                bank_name=bank_name,
                account_no=account_no,
                account_holder_name=holder_name,
                ifsc_code=ifsc,
                branch_name=branch_name,
                initiated_date = timezone.now(),
                status = get_object_or_404(StatusMaster, status_id = 10),
                created_by = user_id
            )
            
            # Call the Cashfree API
            headers = {
                "x-client-id": bank.api_key,
                "x-client-secret": bank.secret_key
            }
            
            payload = {
                "bank_account": account_no,
                "ifsc": ifsc,
                "name": holder_name,
                "branch_name" : branch_name,
                "bank_name" : bank_name,
            }
            
            response = requests.post(bank.url, json=payload, headers=headers)
            
            # Check if the response was successful
            if response.status_code == 200:
                response_data = response.json()  
                
                # Update the BankDetails row with the response data
                employee_update.status = get_object_or_404(StatusMaster, status_id = 11)
                employee_update.utr = response_data.get("utr")
                employee_update.reference_id = response_data.get("reference_id")
                employee_update.completed_date = timezone.now()
                employee_update.updated_by = user_id
                employee_update.save()


                sc_employee_master.objects.filter(
                    employee_id=employee_id, 
                    company_id=company_id
                ).update(
                    bank_name=bank_name,
                    account_no=account_no,
                    account_holder_name = holder_name,
                    ifsc_code=ifsc,
                    branch_name=branch_name
                )

                return JsonResponse(response_data, status=200)
            else:
                response_data = response.json()  # Assuming Cashfree returns JSON data
                
                # Update the BankDetails row with the response data
                employee_update.status = get_object_or_404(StatusMaster, status_id = 11)
                employee_update.reference_id = response_data.get("reference_id")
                employee_update.failure_reason = response_data.get("message")
                employee_update.failed_date = timezone.now()
                employee_update.updated_by = user_id
                employee_update.save()
                return JsonResponse({'message': 'Failed to verify bank details', 'error': response.text}, status=500)
        
        except Exception as e:
            print(e)
            tb = traceback.extract_tb(e.__traceback__)
            cursor.callproc("stp_error_log", [tb[0].name, str(e), request.user.username])
            return JsonResponse({'message': str(e)}, status=500)

@login_required
def check_slot_name(request):
    try:
        # Close previous DB connection if any
        Db.closeConnection()
        m = Db.get_connection()
        cursor = m.cursor()

        if request.method == 'POST':
            # Retrieve the slot name from the POST data
            slot_name = request.POST.get('slot_name')

            # Check if the slot name exists in the database
            if SlotDetails.objects.filter(slot_name=slot_name).exists():
                return JsonResponse({'exists': True})  # Slot name exists
            else:
                return JsonResponse({'exists': False})  # Slot name does not exist
    except Exception as e:
        # Log the exception using your stored procedure
        print(e)
        tb = traceback.extract_tb(e.__traceback__)
        cursor.callproc("stp_error_log", [tb[0].name, str(e), request.user.username])

        # Return an error response
        return JsonResponse({'message': 'An error occurred while processing your request.'}, status=500)
    finally:
        # Always close the cursor and connection to prevent leaks
        cursor.close()
        Db.closeConnection()


def fetch_cities(request):
    try:
        if request.method == "GET":
            state_id = request.GET.get('state_id')
            # Query city names based on the state_id
            city_names = CityMaster.objects.filter(state_id=state_id).values_list('city_id', 'city_name')
            return JsonResponse({'city_names': list(city_names)})
    except Exception as e:
        # Log the exception using your stored procedure
        print(e)

